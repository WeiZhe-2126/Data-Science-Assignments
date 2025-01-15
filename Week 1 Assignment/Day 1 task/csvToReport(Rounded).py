import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Step 1 : Read CSV file
csv_file = 'ST_Omega_Jan23_all.csv'
df = pd.read_csv(csv_file, low_memory=False)

# Ensure correct date format (DD/MM/YY) by using `dayfirst=True`
df['date'] = pd.to_datetime(df['date'], errors='coerce', dayfirst=True)

# Step 2 : Load the Excel file
excel_file = 'ST Omega Saving Report MMM-YY.xlsx'
try:
    book = load_workbook(excel_file)
except FileNotFoundError:
    book = Workbook()

# Step 3 : Select active sheet or a specific sheet
sheet_name = 'Summary'
sheet = book[sheet_name]

# Step 4 : Find the cell with 'Day' and retrieve the dates below it
value_to_find = 'Day'
days_list = []

# Iterate through the rows to find the cell with 'Day'
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == value_to_find:
            # Once we find 'Day', start collecting dates from the cells below it
            next_row = cell.row + 1  # The next row below 'Day'
            # Collect the dates below 'Day' until we reach an empty cell
            while sheet.cell(row=next_row, column=cell.column).value is not None:
                day_value = sheet.cell(row=next_row, column=cell.column).value
                days_list.append(day_value)
                next_row += 1
            break  # Stop after finding the 'Day' and collecting dates

value_to_find_Baseline = 'Baseline'
baseline_value = None

for row in sheet.iter_rows():
    for cell in row:
        if cell.value == value_to_find_Baseline:
            next_column = cell.column + 4
            baseline_value = float(sheet.cell(row=cell.row, column=next_column).value) if isinstance(sheet.cell(row=cell.row, column=next_column).value, (int, float)) else pd.to_numeric(sheet.cell(row=cell.row, column=next_column).value, errors='coerce')
            break
        



# Step 5 : Initialize the row counters for each sum category
next_row_for_HGain = None
next_row_for_CH_KW = None
next_row_for_CHWP_KW = None
next_row_for_CDWP_KW = None
next_row_for_CT_KW = None
next_row_for_Saving_kWh = None
next_row_for_HDR_CH_Eff = None
next_row_for_HDR_Plant_Eff = None
next_row_for_Saving_Percantage = None

# Step 6 : Iterate through each day in the list of days
for current_day in days_list:
    # Ensure the day is in datetime format (DD/MM/YY)
    current_day = pd.to_datetime(current_day, errors='coerce', dayfirst=True)

    if pd.isna(current_day):  # Skip if the day is NaT (Not a Time)
        print(f"Skipping invalid date: {current_day}")
        continue
    
    # Filter the data for the current day
    filtered_data = df[df['date'] == current_day]
    
    if filtered_data.empty:  # If no data is found for the given date
        print(f"No data found for {current_day.strftime('%d/%m/%y')}")
        continue

    # Step 7 : Initialize variables for different sum values
    sum_HGain = None
    rounded_HGain = None
    sum_CH_KW = None
    rounded_CH_KW = None
    sum_CHWP_KW = None
    rounded_CHWP_KW = None
    sum_CDWP_KW = None
    rounded_CDWP_KW = None
    sum_CT_KW = None 
    rounded_CT_KW = None
    sum_Saving_kWh = None 
    rounded_Saving_kWh = None
    sum_HDR_CH_Eff = None 
    rounded_HDR_CH_Eff = None
    sum_HDR_Plant_Eff = None 
    rounded_HDR_Plant_Eff = None
    saving_percentage_value = None 
    rounded_saving_pecentage = None

    value_to_find_Saving_Percantage = 'Energy Saving (%)' # Corresponding Excel header

    # Checking the CSV column and its corresponding Excel header
    if 'HDR_HGain' in df.columns:
        sum_HGain = filtered_data['HDR_HGain'].sum() / 60  # Sum for Total Heat Gain
        value_to_find_HGain = 'Total Heat Gain (RTh)'  # Corresponding Excel header
    if 'HDR_CH_kW' in df.columns:
        sum_CH_KW = filtered_data['HDR_CH_kW'].sum() / 60  # Sum for Chiller Total Energy
        value_to_find_CH_KW = 'Chiller Total Energy (kWh)'  # Corresponding Excel header
    if 'HDR_CHWP_kW' in df.columns:
        sum_CHWP_KW = filtered_data['HDR_CHWP_kW'].sum() / 60  # Sum for CHWP Total Energy
        value_to_find_CHWP_KW = 'CHWP Total Energy (kWh)'  # Corresponding Excel header
    if 'HDR_CDWP_kW' in df.columns:
        sum_CDWP_KW = filtered_data['HDR_CDWP_kW'].sum() / 60  # Sum for CDWP Total Energy
        value_to_find_CDWP_KW = 'CDWP Total Energy (kWh)'  # Corresponding Excel header
    if 'HDR_CT_kW' in df.columns:
        sum_CT_KW = filtered_data['HDR_CT_kW'].sum() / 60  # Sum for CT Total Energy
        value_to_find_CT_KW = 'CT Total Energy (kWh)'  # Corresponding Excel header
    if 'Saving kWh' in df.columns:
        sum_Saving_kWh = filtered_data['Saving kWh'].sum() / 60  # Sum for CT Total Energy
        value_to_find_Saving_kWh = 'Energy Saving (kWh)'  # Corresponding Excel header
    if 'HDR_CH_Eff' in df.columns:
        sum_HDR_CH_Eff = filtered_data['HDR_CH_Eff'].sum()  # Sum for CT Total Energy
        value_to_find_HDR_CH_Eff = 'Chiller Efficiency (kW/RT)'  # Corresponding Excel header
    if 'HDR_Plant_Eff' in df.columns:
        sum_HDR_Plant_Eff = filtered_data['HDR_Plant_Eff'].sum()  # Sum for CT Total Energy
        value_to_find_HDR_Plant_Eff = 'Total Plant Efficiency (kW/RT)'  # Corresponding Excel header

    # Step 8 : Find the specific cell with the Excel header and initialize next_row_for_sums if it's None
    for row in sheet.iter_rows():
        for cell in row:
            # If a sum value was calculated for this column, write it below the found header
            if value_to_find_HGain and cell.value == value_to_find_HGain:
                if next_row_for_HGain is None:
                    next_row_for_HGain = cell.row + 1  # Set next_row_for_HGain to the row below the Excel header
                sheet.cell(row=next_row_for_HGain, column=cell.column).value = rounded_HGain = round(sum_HGain)
                next_row_for_HGain += 1  # Increment to the next row for subsequent sums

            elif value_to_find_CH_KW and cell.value == value_to_find_CH_KW:
                if next_row_for_CH_KW is None:
                    next_row_for_CH_KW = cell.row + 1  # Set next_row_for_CH_KW to the row below the Excel header
                sheet.cell(row=next_row_for_CH_KW, column=cell.column).value = rounded_CH_KW = round(sum_CH_KW)
                next_row_for_CH_KW += 1  # Increment to the next row for subsequent sums

            elif value_to_find_CHWP_KW and cell.value == value_to_find_CHWP_KW:
                if next_row_for_CHWP_KW is None:
                    next_row_for_CHWP_KW = cell.row + 1  # Set next_row_for_CHWP_KW to the row below the Excel header
                sheet.cell(row=next_row_for_CHWP_KW, column=cell.column).value = rounded_CHWP_KW = round(sum_CHWP_KW)
                next_row_for_CHWP_KW += 1  # Increment to the next row for subsequent sums

            elif value_to_find_CDWP_KW and cell.value == value_to_find_CDWP_KW:
                if next_row_for_CDWP_KW is None:
                    next_row_for_CDWP_KW = cell.row + 1  # Set next_row_for_CDWP_KW to the row below the Excel header
                sheet.cell(row=next_row_for_CDWP_KW, column=cell.column).value = rounded_CDWP_KW = round(sum_CDWP_KW)
                next_row_for_CDWP_KW += 1  # Increment to the next row for subsequent sums

            elif value_to_find_CT_KW and cell.value == value_to_find_CT_KW:
                if next_row_for_CT_KW is None:
                    next_row_for_CT_KW = cell.row + 1  # Set next_row_for_CT_KW to the row below the Excel header
                sheet.cell(row=next_row_for_CT_KW, column=cell.column).value = rounded_CT_KW = round(sum_CT_KW)
                next_row_for_CT_KW += 1  # Increment to the next row for subsequent sums

            elif value_to_find_Saving_kWh and cell.value == value_to_find_Saving_kWh:
                if next_row_for_Saving_kWh is None:
                    next_row_for_Saving_kWh = cell.row + 1  # Set next_row_for_CT_KW to the row below the Excel header
                sheet.cell(row=next_row_for_Saving_kWh, column=cell.column).value = rounded_Saving_kWh = round(sum_Saving_kWh)
                next_row_for_Saving_kWh += 1  # Increment to the next row for subsequent sums

            elif value_to_find_HDR_CH_Eff and cell.value == value_to_find_HDR_CH_Eff:
                if next_row_for_HDR_CH_Eff is None:
                    next_row_for_HDR_CH_Eff = cell.row + 1  # Set next_row_for_CT_KW to the row below the Excel header
                sum_HDR_CH_Eff = (sum_CH_KW/sum_HGain)
                sheet.cell(row=next_row_for_HDR_CH_Eff, column=cell.column).value = rounded_HDR_CH_Eff = round(sum_HDR_CH_Eff, 3)
                next_row_for_HDR_CH_Eff += 1  # Increment to the next row for subsequent sums

            elif value_to_find_HDR_Plant_Eff and cell.value == value_to_find_HDR_Plant_Eff:
                if next_row_for_HDR_Plant_Eff is None:
                    next_row_for_HDR_Plant_Eff = cell.row + 1  # Set next_row_for_CT_KW to the row below the Excel header
                sum_HDR_Plant_Eff = ( (sum_CH_KW + sum_CHWP_KW + sum_CDWP_KW + sum_CT_KW ) / sum_HGain )
                sheet.cell(row=next_row_for_HDR_Plant_Eff, column=cell.column).value = rounded_HDR_Plant_Eff = round( sum_HDR_Plant_Eff, 3 )
                next_row_for_HDR_Plant_Eff += 1  # Increment to the next row for subsequent sums

            elif value_to_find_Saving_Percantage and cell.value == value_to_find_Saving_Percantage:
                if next_row_for_Saving_Percantage is None:
                    next_row_for_Saving_Percantage = cell.row + 1
                saving_percentage_value = (rounded_Saving_kWh / (rounded_HGain * baseline_value)) * 100.00
                sheet.cell(row=next_row_for_Saving_Percantage, column=cell.column).value = rounded_saving_pecentage = round(saving_percentage_value, 1)
                next_row_for_Saving_Percantage += 1

    # Print to verify the output for each day
    print(f"Sum for {current_day.strftime('%d/%m/%y')}:")
    if rounded_HGain is not None:
        print(f" - Total Heat Gain (RTh): {rounded_HGain}")
    if rounded_CH_KW is not None:
        print(f" - Chiller Total Energy (kWh): {rounded_CH_KW}")
    if rounded_CHWP_KW is not None:
        print(f" - CHWP Total Energy (kWh): {rounded_CHWP_KW}")
    if rounded_CDWP_KW is not None:
        print(f" - CDWP Total Energy (kWh): {rounded_CDWP_KW}")
    if rounded_CT_KW is not None:
        print(f" - CT Total Energy (kWh): {rounded_CT_KW}")
    if rounded_Saving_kWh is not None:
        print(f" - Energy Saving (kWh): {rounded_Saving_kWh}")
    if rounded_HDR_CH_Eff is not None:
        print(f" - Chiller Efficiency (kW/RT): {rounded_HDR_CH_Eff}")
    if rounded_HDR_Plant_Eff is not None:
        print(f" - Total Plant Efficiency (kW/RT): {rounded_HDR_Plant_Eff}")
    if rounded_saving_pecentage is not None:
        print(f" - Energy Saving (%): {rounded_saving_pecentage}")

# Step 9 : Save the workbook
book.save(excel_file)
print(f"Excel file updated successfully with sums for the listed days.")
